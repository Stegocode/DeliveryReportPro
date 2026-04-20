# -*- coding: utf-8 -*-
"""
financial_generator_app.py
===========================
DeliveryReportPro report generator — reads scraped exports, calculates
per-stop delivery costs, assembles the multi-bucket financial report,
and writes the formatted .xlsx.
Credentials come from os.environ (injected by pipeline.py).
Progress is emitted via ProgressEmitter; errors via ErrorRegistry.
Output goes to %APPDATA%\\DeliveryReportPro\\exports\\ — no admin rights needed.
"""

import os
import re
import time
import requests
import pandas as pd
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Credentials from environment (set by pipeline.py, cleared after run)
ORS_API_KEY    = os.getenv("ORS_API_KEY", "")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
MONDAY_TOKEN   = os.getenv("MONDAY_API_TOKEN", "")

# Output to APPDATA — no admin rights needed
EXPORT_BASE = os.path.join(
    os.environ.get("APPDATA", os.path.expanduser("~")),
    "DeliveryReportPro", "exports"
)
WAREHOUSE = "4600 NW St Helens Rd, Portland, OR 97210"

# ── Monday DELIVERY SCHEDULER ──────────────────────────────────────
DELIVERY_SCHED_BOARD = "8639744112"
CRATE_COL_ID         = "color_mkns94j8"
MONDAY_API_URL       = "https://api.monday.com/v2"

_CRATE_LABEL_MAP = {
    'IN-BOX':                  'IN-BOX',
    'OUT OF BOX':              'OUT OF BOX',
    'OUT OF BOX + INST':       'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTA':      'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTAL':     'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTALL':    'OUT OF BOX + INSTALL',
    'OUT OF BOX + INSTALL(S)': 'OUT OF BOX + INSTALL',
}

# ── Truck sets ─────────────────────────────────────────────────────
OWN_FLEET_TRUCKS  = {"56","58","62","64","68","72","S.VAN","SVAN"}
HUB_PREFIX    = "HUB"
THIRD_PARTY   = {"BRIANK","J REED","TB"}
WILL_CALL_TYPES = {"Pickup","Drop Ship","Transfer","Unknown"}

# ── Cost constants ─────────────────────────────────────────────────
X001_COST          = 20.00
CRATED_PER_PIECE   = 18.70
UNCRATED_PER_PIECE = 28.00

# ── Piece cost — piecewise logarithmic curve ───────────────────────
# Labor is sub-linear at low piece counts (fixed crew overhead dominates)
# and settles to a flat per-piece rate for large multifamily stops.
#
# Anchor points derived from observed crew times:
#   1 piece  — 2-man crew, ~17.5 min avg  →  $14.58
#   10 pieces — avg of 2-man×1.75hr / 3-man×1.25hr  →  $90.62
#   200 pieces — 9-man crew, 8 hrs  →  $1,800.00
#
# 1–10 pieces: power curve  cost = BASE_1PC × pieces^0.7934
# 10+ pieces:  linear       cost = AT_10PC  + $9.00 × (pieces − 10)
#
# Crated follows the same curve scaled by the crated/uncrated rate ratio.

import math as _math

_ACTUAL_RATE     = 22.85   # actual avg delivery crew rate — Q1 2026 payroll ($87,016 reg / 3,808 hrs)
_OVERHEAD_RATE   = 12.15   # overhead per piece (maintenance, insurance, WC, claims — see README)
_TIME_CREW_EXP   = 0.832   # time-scaling exponent: derived from 2-man=1.75hr, 3-man=1.25hr at 10pcs
                            # time_at_10(N) = 1.75 × (2/N)^0.832
_BASE_1PC_U      = 2 * (17.5 / 60) * _ACTUAL_RATE   # 1 piece — always 2-man, 17.5 min avg
_AT_200PC_U      = 9 * 8   * _ACTUAL_RATE            # 200 pieces — 9-man, 8 hrs (multifamily anchor)
_CRATED_SCALE    = CRATED_PER_PIECE / UNCRATED_PER_PIECE   # 0.668


def _crew_at10(crew):
    """Labor cost at 10 pieces for a given crew size."""
    n            = max(crew, 2)
    time_at_10   = 1.75 * (2 / n) ** _TIME_CREW_EXP
    return round(n * time_at_10 * _ACTUAL_RATE, 4)


def piece_cost(pieces, crate_key='OUT OF BOX', crew=2):
    """
    Returns the all-in labor + overhead cost for uncrating and staging pieces.

    Labor curve (piecewise, crew-adjusted):
      1–10 pieces:  power curve  cost = BASE_1PC × pieces^exp
                    where exp = log(at_10(crew) / BASE_1PC) / log(10)
      10+ pieces:   linear       cost = at_10(crew) + marginal × (pieces − 10)
                    where marginal = (AT_200 − at_10(crew)) / 190

    Overhead: $12.15/piece flat (maintenance, insurance, WC, claims).
    Crated stops scaled by 0.668 (crated/uncrated rate ratio).

    Calibration (Q1 2026):
      Avg delivery rate: $22.85/hr (reg only)
      1 piece:    2-man, 17.5 min
      10 pieces:  2-man=1.75hr, 3-man=1.25hr  (crew-adjusted)
      200 pieces: 9-man, 8 hrs
    """
    if pieces <= 0:
        return 0.0

    at_10  = _crew_at10(crew)
    marg   = (_AT_200PC_U - at_10) / (200 - 10)
    exp    = _math.log(at_10 / _BASE_1PC_U) / _math.log(10)

    if pieces <= 10:
        labor = _BASE_1PC_U * (pieces ** exp)
    else:
        labor = at_10 + marg * (pieces - 10)

    total = labor + _OVERHEAD_RATE * pieces
    if crate_key == 'IN-BOX':
        total *= _CRATED_SCALE
    return round(total, 2)
INSURANCE_PER_TRUCK= 20.00
MF_THRESHOLD       = 33.00
FLOOR_CHARGE       = 18.70
FUEL_SHORT         = 6.60
FUEL_LONG          = 9.90
MILEAGE_CHARGE     = 30.00
MILEAGE_FREE_MAX   = 30
MILEAGE_MAX        = 125
MPG_DIESEL         = 6.0
LABOR_RATE         = 25.00

# ── Formatting ─────────────────────────────────────────────────────
BLUE   = "BDD7EE"; ORANGE = "FCE4D6"; GREEN = "E2EFDA"; GRAY = "D9D9D9"
DOLLAR_FMT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
THIN   = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BUCKET_FILL = {
    "Delivery":         None,
    "Will Call":        PatternFill("solid", fgColor="EBF1DE"),
    "Storage Release":  PatternFill("solid", fgColor="FFF2CC"),
    "RMA":              PatternFill("solid", fgColor="E2D5F0"),
}

def bold(sz=11):   return Font(name="Calibri", bold=True,  size=sz)
def normal(sz=11): return Font(name="Calibri", bold=False, size=sz)

def fmt_cell(c, fmt, bg=None):
    c.border = BORDER; c.font = normal()
    c.alignment = Alignment(horizontal="right" if fmt in ("dollar","pct","int") else "left")
    if fmt == "dollar": c.number_format = DOLLAR_FMT
    elif fmt == "pct":  c.number_format = '0.0'
    if bg: c.fill = PatternFill("solid", fgColor=bg)

def write_headers(ws, headers, row=1):
    for ci, (lbl, w) in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=lbl)
        c.font = bold(); c.border = BORDER
        c.fill = PatternFill("solid", fgColor=GRAY)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w


# ── X-code cost table ──────────────────────────────────────────────
XCODE_COST = {
    'X001':  20.00, 'X015':  51.00, 'X018':  33.00,
    'B002':  50.00,
    'X021':  10.20, 'X022':  20.40, 'X029':   5.50,
    'X100':  19.00, 'X103A': 27.50, 'X103B': 27.50, 'X103': 27.50,
    'X151':  37.50, 'X152':   3.00, 'X154':  80.00, 'X156': 275.00,
    'X157': 145.00, 'X159':   6.00, 'X164':  87.00, 'X166':  12.00,
    'X201': 124.50, 'X204': 140.50, 'X208':  75.00, 'X212': 141.00,
    'X214':  20.00, 'X216': 197.00, 'X222': 210.00, 'X224': 231.00,
    'X226':  32.00, 'X230':   5.50, 'X232': 130.00, 'X238': 346.50,
    'X251':  12.50, 'X253':  12.50, 'X255':  15.50,
    'X257':  25.00, 'X259':  25.00, 'X260':  27.50, 'X261':  27.50,
    'X301':  31.00, 'X303':  18.50, 'X328':  19.00, 'X336':  27.50,
}

# ── Exclude / parts / model detection ─────────────────────────────
EXCLUDE_MODELS = {
    'CC FEES','CC FEE','ACCOMM','MGMT ACCOMMODATION',
    'LIEN','CAT TAX','NSF','CHARGEBACK CC','PAYROLL','SM WO',
    'ACCTG - WARRANTIES','ACCOUNTING','REFUND','WF FEES','TD','FACTORY',
    'STORAGE','RESTOCK','LATE FEE','99PRICEADJ','99MISC50',
    'CUSTOMERS HOME','DAMAGE BY DELIVERY','FREIGHT','CREDIT CARD FEE',
    'CONVERSION-GAS','RETURN','MEMO','SPECIAL',
}

# Accessories — have sale prices, not physical crate/uncrate pieces
ACCESSORY_MODELS = {
    'DW INSTALL KIT','DW KIT','LAUNDRYPACK','LAUNDRYPACK-GAS','LAUNDRYPACK-ELECT',
    'WATERLINE','WATERHOSE SS','WATERHOSE RUB','GASLINE','DRYERCORD',
    'RANGECORD','DWELBOW','BRACKET','STEAM DRYER','STEAM DRYER ',
    'WASHERIN','52525','3PRONGCORD','110 POWERCORD','DW PANS','LAUNDRY PAN',
    'REFERINSTALL',
}

# Stair/oversized surcharges — sale price on invoice, cost = 3/5 of sale, paid to crew
# B003 is always $50. STAIR codes are $50 increments.
STAIR_COST_RATIO = 3 / 5
STAIR_MODELS = {'B003', 'STAIR 5-10', 'STAIR 11-15', 'STAIR 16-20', 'STAIR 21-25'}

# Same logic as monday_populator.parse_orders / is_model()
# Accessories and stair codes are excluded from piece counts but NOT from revenue
NON_MODEL_CODES = EXCLUDE_MODELS | ACCESSORY_MODELS | STAIR_MODELS | {
    'X001','X002','X003','X004','X007','X011','X012','X013','X014','X015',
    'X018','X020','X021','X022','X023','X025','X028','X029','X075','X100',
    'X101','X102','X103','X103A','X103B','X151','X152','X154','X156','X157',
    'X159','X164','X166','X201','X204','X208','X210','X212','X214','X216',
    'X222','X224','X226','X230','X232','X238','X251','X253','X255','X257',
    'X259','X260','X261','X301','X303','X328','X336','X998','MEMO',
    'ADA DW PANS 18"','CONVERSION-DOORSWING',
    'TPI','CONVERSION-GAS','DRYER INSTALL','REDEL','B002',
}

def is_model(model_key):
    return bool(model_key) and model_key.upper() not in NON_MODEL_CODES

def is_service(model):
    m = str(model or '').strip().upper()
    return (m.startswith('X') and len(m) <= 5) or m in {'TPI','SPECIAL','DELIVERY-DELUXE','B002'}

def is_accessory(model):
    """Accessories: physical add-ons with sale price but not crate/uncrate pieces."""
    return str(model or '').strip().upper() in ACCESSORY_MODELS

def is_stair(model):
    """Stair/oversized surcharges: sale price on invoice, cost = 3/5 of sale."""
    return str(model or '').strip().upper() in STAIR_MODELS

def normalize_truck(raw):
    t = str(raw or '').strip().upper()
    if t.startswith('TRUCK '): return t.replace('TRUCK ', '').strip()
    if t.startswith('HUB #'):  return 'HUB ' + t.replace('HUB #','').strip().zfill(2)
    if t.startswith('HUB'):
        # e.g. "HUB1", "HUB 1", "HUB04"
        num = t.replace('HUB','').strip().lstrip('#').strip()
        if num.isdigit():
            return 'HUB ' + num.zfill(2)
    return t

# Normalize truck names from route sheet PDF
# "6" → "62", "H2B 01" → "HUB 01" etc.
_ROUTE_TRUCK_MAP = {
    '6': '62',
    'H2B 01': 'HUB 01', 'H2B 02': 'HUB 02', 'H2B 03': 'HUB 03',
    'HUB1': 'HUB 01',   'HUB2': 'HUB 02',   'HUB3': 'HUB 03',
}

def normalize_route_truck(raw):
    t = str(raw or '').strip()
    if t in _ROUTE_TRUCK_MAP:
        return _ROUTE_TRUCK_MAP[t]
    # "HUB #1" style
    if t.upper().startswith('HUB #'):
        return 'HUB ' + t.upper().replace('HUB #','').zfill(2)
    return t.upper()

def categorize(truck, dtype):
    t = str(truck or '').strip().upper()
    d = str(dtype or '').strip()
    if t in {'RETURN','RETURNS'}:                            return 'Return'
    if t == 'UNPAID':                                        return 'Unpaid'
    if t in {x.upper() for x in THIRD_PARTY}:               return 'Will Call'
    if t == 'OWN':                                         return 'Will Call'
    if d in WILL_CALL_TYPES:                                 return 'Will Call'
    if not t or t in {'STORAGE','UPS','FEDEX','BK CROW',
                      'BKCROW','TRADE B','TRADE BK'}:        return 'Will Call'
    return 'Delivery'

def is_own_truck(truck):
    return truck.upper() in {x.upper() for x in OWN_FLEET_TRUCKS}

def is_hub_truck(truck):
    return truck.upper().startswith(HUB_PREFIX)

TRUCK_SORT = {
    '56':1,'58':2,'62':3,'64':4,'68':5,'72':6,'S.VAN':7,'SVAN':7,
    'OWN':8,'HUB 01':9,'HUB 02':10,'HUB 03':11
}


# ── Route sheet PDF parser ─────────────────────────────────────────

def parse_route_sheet(pdf_path, emitter=None):
    """
    Parse route sheet PDF → {truck: [(stop_num, order_id, address), ...]}
    Uses phone number as anchor to extract clean addresses.
    Handles multi-line customer names and pre-buffered address lines.
    """
    import pdfplumber as _pdfplumber
    SKIP = {'BEND SHOWROOM TRANSFER', 'PORTLAND SHOWROOM TRANSFER'}
    _TRUCK_MAP = {'6':'62','H2B 01':'HUB 01','H2B 02':'HUB 02','H2B 03':'HUB 03'}

    def norm_truck(raw):
        t = str(raw or '').strip()
        if t in _TRUCK_MAP: return _TRUCK_MAP[t]
        if re.match(r'^HUB\s*#?\s*\d+$', t, re.I):
            return 'HUB ' + re.search(r'\d+', t).group().zfill(2)
        return t.upper()

    def extract_address(text):
        """Find Street, City STATE Zip pattern and return it."""
        sz = re.search(
            r'(.+?,\s*[\w\s.]+(?:OR|WA|CA|ID|NV|AZ)\s+\d{5}(?:-\d{4})?)',
            text, re.IGNORECASE
        )
        if not sz:
            return ''
        return re.sub(r'(\d{5})-\d{4}', r'\1', sz.group(1)).strip()

    stop_order          = defaultdict(list)
    current_truck       = None
    current_page_num    = 1
    current_total_pages = 1

    PHONE_RE = re.compile(r'\(\d{3}\)\s*\d{3}-\d{4}')
    STOP_RE  = re.compile(r'^(\d+)\s+(\d{5})\s+(.*)')

    with _pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text  = page.extract_text() or ''
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            if not lines: continue

            is_cont = current_truck and current_page_num < current_total_pages
            if not is_cont:
                current_truck = norm_truck(lines[0])

            for line in reversed(lines):
                m = re.search(r'Page\s+(\d+)\s+of\s+(\d+)', line, re.I)
                if m:
                    current_page_num    = int(m.group(1))
                    current_total_pages = int(m.group(2))
                    break

            in_table     = False
            pending      = None
            pre_addr_buf = ''
            pre_cust_buf = ''

            for line in lines:
                if '#' in line and 'Order' in line and 'Customer' in line:
                    in_table = True
                    continue
                if not in_table: continue
                if re.match(r'\d{2}/\d{2}/\d{4}\s+Page', line): continue

                m = STOP_RE.match(line)
                if m:
                    if pending:
                        stop_order[current_truck].append(pending)
                        pending = None

                    stop_num  = int(m.group(1))
                    order_id  = int(m.group(2))
                    remainder = m.group(3).strip()

                    phone_m = PHONE_RE.search(remainder)
                    if phone_m:
                        before_phone = remainder[:phone_m.start()].strip()
                        if pre_addr_buf:
                            before_phone = pre_addr_buf + ' ' + before_phone
                            pre_addr_buf = ''
                        addr     = extract_address(before_phone)
                        customer = pre_cust_buf.strip() if pre_cust_buf else ''
                        pre_cust_buf = ''
                        if not any(s in (customer+addr).upper() for s in SKIP):
                            stop_order[current_truck].append((stop_num, order_id, customer, addr))
                    else:
                        if pre_addr_buf:
                            addr = pre_addr_buf; pre_addr_buf = ''
                            customer = remainder.strip()
                            pending = (stop_num, order_id, customer, addr)
                        else:
                            pending = (stop_num, order_id, remainder.strip(), '')
                        pre_cust_buf = ''

                elif pending:
                    stop_num, order_id, customer, addr = pending
                    phone_m = PHONE_RE.search(line)
                    if phone_m:
                        before_phone = line[:phone_m.start()].strip()
                        if not addr:
                            addr = re.sub(r'(\d{5})-\d{4}', r'\1', before_phone)
                        else:
                            addr = re.sub(r'(\d{5})-\d{4}', r'\1', addr + ' ' + before_phone)
                        if not any(s in (customer+addr).upper() for s in SKIP):
                            stop_order[current_truck].append((stop_num, order_id, customer, addr))
                        pending = None
                    else:
                        if re.match(r'^\d+$', line):
                            addr = addr.rstrip('-') + line
                            pending = (stop_num, order_id, customer, addr)
                        else:
                            customer = (customer + ' ' + line).strip()
                            pending = (stop_num, order_id, customer, addr)

                else:
                    phone_m = PHONE_RE.search(line)
                    if not phone_m:
                        if re.match(r'^\d+\s+\w', line):
                            pre_addr_buf = re.sub(r'(\d{5})-\d{4}', r'\1', line.rstrip('-'))
                        elif line and not re.match(r'\d{2}/\d{2}/\d{4}', line):
                            pre_cust_buf = line

    if pending:
        stop_order[current_truck].append(pending)

    for truck in stop_order:
        stop_order[truck].sort(key=lambda x: x[0])

    total = sum(len(v) for v in stop_order.values())
    good  = sum(1 for v in stop_order.values() for s in v if s[3])
    if emitter: emitter.log(f'Route sheet parsed: {good}/{total} stops with addresses')
    for truck, stops in sorted(stop_order.items()):
        if emitter: emitter.log(f'  {truck}: {[s[1] for s in stops]}')

    return dict(stop_order)


def get_crate_status(order_ids, emitter=None):
    """
    Query Monday DELIVERY SCHEDULER board for crate status per order.
    Returns {order_id_str: 'IN-BOX' | 'OUT OF BOX' | 'OUT OF BOX + INSTALL'}
    """
    if not MONDAY_TOKEN:
        if emitter: emitter.log("MONDAY_API_TOKEN not set — defaulting OUT OF BOX", "warn")
        return {}

    target = {str(oid) for oid in order_ids}
    result = {}
    headers = {"Authorization": MONDAY_TOKEN, "Content-Type": "application/json"}
    cursor  = None

    if emitter: emitter.log("Fetching crate status from Monday...")

    while True:
        if cursor:
            query = """
            query ($boardId: [ID!], $cursor: String!) {
              boards(ids: $boardId) {
                items_page(limit: 500, cursor: $cursor) {
                  cursor
                  items {
                    name
                    column_values(ids: ["%s"]) { id text value }
                  }
                }
              }
            }
            """ % CRATE_COL_ID
            variables = {"boardId": DELIVERY_SCHED_BOARD, "cursor": cursor}
        else:
            query = """
            query ($boardId: [ID!]) {
              boards(ids: $boardId) {
                items_page(limit: 500) {
                  cursor
                  items {
                    name
                    column_values(ids: ["%s"]) { id text value }
                  }
                }
              }
            }
            """ % CRATE_COL_ID
            variables = {"boardId": DELIVERY_SCHED_BOARD}

        try:
            r = requests.post(
                MONDAY_API_URL,
                json={"query": query, "variables": variables},
                headers=headers, timeout=15
            )
            data = r.json()
        except Exception as e:
            if emitter: emitter.log(f"Monday API error: {e}", "warn")
            break

        if "errors" in data:
            if emitter: emitter.log(f"Monday errors: {data['errors']}", "warn")
            break

        page  = data["data"]["boards"][0]["items_page"]
        items = page.get("items", [])
        cursor = page.get("cursor")

        for item in items:
            # Extract order number — first numeric token in item name
            name_match = re.search(r'\b(\d{4,6})\b', item["name"])
            if not name_match:
                continue
            order_num = name_match.group(1)
            if order_num not in target:
                continue

            for col in item["column_values"]:
                if col["id"] == CRATE_COL_ID:
                    raw_label = (col.get("text") or "").strip().upper()
                    # Match against known label prefixes
                    matched = None
                    for key, canonical in _CRATE_LABEL_MAP.items():
                        if raw_label.startswith(key):
                            matched = canonical
                            break
                    if matched:
                        result[order_num] = matched
                    break

        if not cursor:
            break

    found    = len(result)
    missing  = [str(o) for o in order_ids if str(o) not in result]
    if emitter: emitter.log(f"Crate status: {found} found, {len(missing)} missing")
    if missing:
        if emitter: emitter.log(f"Missing crate status (defaulting OUT OF BOX): {missing}", "warn")

    return result


# ── Diesel price ───────────────────────────────────────────────────

def get_diesel_price(emitter=None, errors=None, prompt_answers=None):
    """
    Three-layer diesel price lookup:
      1. EIA API (federal data, most reliable)
      2. Google Custom Search fallback
      3. UI prompt if both fail

    Args:
        emitter:        ProgressEmitter for logging
        errors:         ErrorRegistry for warnings
        prompt_answers: dict — will contain 'diesel_price' if UI answered

    Returns:
        float diesel price per gallon
    """
    def log(msg, level='info'):
        if emitter: emitter.log(msg, level)

    # Layer 1: EIA API
    try:
        r = requests.get(
            "https://api.eia.gov/v2/petroleum/pri/gnd/data/"
            "?api_key=DEMO&frequency=weekly&data[0]=value"
            "&facets[product][]=DU&facets[duoarea][]=NUS"
            "&sort[0][column]=period&sort[0][direction]=desc&offset=0&length=1",
            timeout=8
        )
        val = float(r.json()['response']['data'][0]['value'])
        log(f'Diesel price (EIA): ${val:.3f}/gal', 'good')
        if emitter: emitter.done('diesel', f'Diesel: ${val:.3f}/gal (EIA live data)')
        return val
    except Exception as e:
        log(f'EIA API unavailable: {e}', 'warn')

    # Layer 2: Google Custom Search
    if GOOGLE_API_KEY:
        try:
            r = requests.get(
                "https://www.googleapis.com/customsearch/v1",
                params={
                    'key': GOOGLE_API_KEY,
                    'cx':  '017576662512468239146:omuauf_lfve',
                    'q':   'current US average diesel price per gallon',
                    'num': 1,
                },
                timeout=8
            )
            snippet = r.json().get('items', [{}])[0].get('snippet', '')
            # Look for a price pattern like $3.87 or 3.879
            m = re.search(r'\$?\s*(\d+\.\d{2,3})', snippet)
            if m:
                val = float(m.group(1))
                if 2.0 < val < 8.0:   # sanity check
                    log(f'Diesel price (Google): ${val:.3f}/gal', 'good')
                    if emitter: emitter.done('diesel', f'Diesel: ${val:.3f}/gal (Google)')
                    return val
        except Exception as e:
            log(f'Google diesel search failed: {e}', 'warn')

    # Layer 3: UI prompt
    if errors:
        errors.warn('diesel', 'DIESEL_MANUAL',
                    'Could not fetch diesel price automatically — manual entry required')
    if emitter:
        emitter.warn('diesel', 'Enter diesel price manually')
        emitter.prompt('diesel_price', {
            'label':    'Diesel Price',
            'question': "What's today's diesel price?",
            'sub':      'Check GasBuddy or your fleet card — enter price per gallon',
        })

    # Wait for UI answer
    if prompt_answers is not None:
        import threading, time as _time
        for _ in range(300):   # up to 5 minutes
            if 'diesel_price' in prompt_answers:
                val = float(prompt_answers['diesel_price'])
                log(f'Diesel price (manual): ${val:.3f}/gal')
                if emitter: emitter.done('diesel', f'Diesel: ${val:.3f}/gal (manual)')
                return val
            _time.sleep(1)

    # Last resort fallback
    log('Using diesel fallback price $3.80/gal', 'warn')
    return 3.80


# ── ORS helpers ────────────────────────────────────────────────────

def sanitize_address(address):
    """Clean a full assembled address string for ORS geocoding.
    Handles: semicolons in street (keep only part before semicolon),
    double spaces, ZIP+4 format.
    Called on the already-assembled 'Street, City, State Zip' string.
    """
    import re as _re
    a = str(address or '').strip()
    # Split into street vs rest on first comma
    parts = a.split(',', 1)
    street = parts[0]
    rest   = parts[1] if len(parts) > 1 else ''
    # Strip semicolon and after from street only
    street = street.split(';')[0].strip()
    # Reassemble
    a = (street + ', ' + rest.strip()) if rest.strip() else street
    # Collapse double spaces
    a = _re.sub(r' {2,}', ' ', a)
    # Strip ZIP+4 to 5 digits only
    a = _re.sub(r'(\d{5})-\d{4}', r'\1', a)
    return a.strip()



# ── Geocode cache ──────────────────────────────────────────────────
_geocode_cache = {}

def geocode(address, emitter=None):
    """Geocode using Google Maps API. Returns [lon, lat] to match ORS format.
    Results are cached in memory to avoid duplicate API calls."""
    if not GOOGLE_API_KEY: return None
    clean = sanitize_address(address)
    if clean in _geocode_cache:
        return _geocode_cache[clean]
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/geocode/json",
            params={'address': clean, 'key': GOOGLE_API_KEY},
            timeout=15
        )
        data = r.json()
        if data.get('status') != 'OK':
            if emitter: emitter.log(f"Geocode failed '{clean}': {data.get('status')}", 'warn')
            _geocode_cache[clean] = None
            return None
        loc = data['results'][0]['geometry']['location']
        lat, lng = loc['lat'], loc['lng']
        if not (-130 < lng < -60 and 24 < lat < 50):
            if emitter: emitter.log(f"Out-of-range coords for '{clean}': {lat},{lng}", 'warn')
            _geocode_cache[clean] = None
            return None
        result = [lng, lat]
        _geocode_cache[clean] = result
        return result
    except Exception as e:
        if emitter: emitter.log(f"Geocode failed '{clean}': {e}", 'warn')
    _geocode_cache[clean] = None
    return None


def drive_miles(origin, dest, emitter=None):
    """Driving distance in miles via Google Distance Matrix API."""
    if not GOOGLE_API_KEY: return None
    try:
        origin_str = f"{origin[1]},{origin[0]}"
        dest_str   = f"{dest[1]},{dest[0]}"
        r = requests.get(
            "https://maps.googleapis.com/maps/api/distancematrix/json",
            params={
                'origins':      origin_str,
                'destinations': dest_str,
                'mode':         'driving',
                'units':        'imperial',
                'key':          GOOGLE_API_KEY,
            },
            timeout=15
        )
        data = r.json()
        if data.get('status') != 'OK':
            if emitter: emitter.log(f"Distance Matrix failed: {data.get('status')}", 'warn')
            return None
        element = data['rows'][0]['elements'][0]
        if element.get('status') != 'OK':
            if emitter: emitter.log(f"Distance element failed: {element.get('status')}", 'warn')
            return None
        meters = element['distance']['value']
        return round(meters / 1609.34, 1)
    except Exception as e:
        if emitter: emitter.log(f"Distance failed: {e}", 'warn')
    return None


def hub_mileage_charges(address, order_id, name, warehouse_coords, emitter=None):
    if not warehouse_coords: return None, 0.0, 0.0
    dest = geocode(address, emitter)
    time.sleep(0.05)
    if not dest:
        if emitter: emitter.log(f'Order {order_id} ({name}): geocode failed', 'warn')
        return None, 0.0, 0.0
    miles = drive_miles(warehouse_coords, dest, emitter)
    time.sleep(0.05)
    if miles is None or miles > 200:
        if emitter: emitter.log(f'Order {order_id} ({name}): distance failed', 'warn')
        return None, 0.0, 0.0
    if miles < MILEAGE_FREE_MAX:
        if emitter: emitter.log(f'Order {order_id} ({name}): {miles}mi → fuel ${FUEL_SHORT}')
        return miles, 0.0, FUEL_SHORT
    elif miles <= MILEAGE_MAX:
        if emitter: emitter.log(f'Order {order_id} ({name}): {miles}mi → fuel ${FUEL_LONG} + mileage ${MILEAGE_CHARGE}')
        return miles, MILEAGE_CHARGE, FUEL_LONG
    else:
        if emitter: emitter.log(f'Order {order_id} ({name}): {miles}mi — over {MILEAGE_MAX}mi, flagged', 'warn')
        return miles, 0.0, FUEL_LONG


def own_route_fuel(stops_in_order, diesel_price, emitter=None):
    """
    Chained route: warehouse → stop1 → stop2 → ... → warehouse.
    stops_in_order: [(order_id, name, address), ...] already in route order.
    Returns {order_id: fuel_cost}
    """
    if not stops_in_order or not ORS_API_KEY:
        return {oid: 0.0 for oid, _, _ in stops_in_order}

    warehouse_coords = geocode(WAREHOUSE, emitter)
    if not warehouse_coords:
        if emitter: emitter.log('Could not geocode warehouse', 'warn')
        return {oid: 0.0 for oid, _, _ in stops_in_order}

    chain = [warehouse_coords]
    valid = []
    for oid, name, addr in stops_in_order:
        if emitter: emitter.log(f'  Geocoding {name}…')
        coords = geocode(addr, emitter)
        time.sleep(0.05)
        if coords:
            chain.append(coords)
            valid.append((oid, name))
        else:
            if emitter: emitter.log(f'Order {oid} ({name}): geocode failed — $0 fuel', 'warn')

    chain.append(warehouse_coords)

    leg_miles   = []
    total_miles = 0.0
    for i in range(len(chain) - 1):
        m = drive_miles(chain[i], chain[i+1], emitter) or 0.0
        time.sleep(0.05)
        leg_miles.append(m)
        total_miles += m

    total_fuel = (total_miles / MPG_DIESEL) * diesel_price
    if emitter: emitter.log(f'Route: {total_miles:.1f}mi total → ${total_fuel:.2f} fuel')

    n = len(valid)
    if n == 0:
        return {}

    return_share = ((leg_miles[-1] if leg_miles else 0.0) / MPG_DIESEL * diesel_price) / n
    fuel_per_stop = {}
    for i, (oid, name) in enumerate(valid):
        inbound   = leg_miles[i] if i < len(leg_miles) - 1 else 0.0
        stop_fuel = round((inbound / MPG_DIESEL * diesel_price) + return_share, 2)
        fuel_per_stop[oid] = stop_fuel
        if emitter: emitter.log(f'  Order {oid} ({name}): {inbound:.1f}mi → ${stop_fuel:.2f}')

    return fuel_per_stop


# ── Prompt helpers ─────────────────────────────────────────────────

def prompt_int(prompt_id, context, prompt_answers, emitter, allow_zero=True):
    import time as _time

    emitter.prompt(prompt_id, context)

    for _ in range(300):
        if prompt_id in prompt_answers:
            raw = prompt_answers[prompt_id]
            try:
                val = int(float(str(raw)))  # handle int, float, or string
                if not allow_zero and val < 1:
                    del prompt_answers[prompt_id]
                    emitter.prompt(prompt_id, {**context,
                        'sub': 'Must be at least 1 — please re-enter'})
                    continue
                emitter.log(f'  Prompt {prompt_id} answered: {val}')
                return val
            except (ValueError, TypeError):
                emitter.log(f'  Prompt {prompt_id} bad value: {raw!r} — retrying', 'warn')
                del prompt_answers[prompt_id]
                continue
        _time.sleep(0.5)

    emitter.log(f'Prompt {prompt_id} timed out — using default 1', 'warn')
    return 1


def is_multifamily_customer(billing_customer_type):
    """Matches flat_file_generator_v5 logic — Billing Customer Type contains MULTI."""
    return 'MULTI' in str(billing_customer_type or '').upper()


# ── Load scraped files ─────────────────────────────────────────────

def load_files(bulk_path, serial_path, orders_path, target_date, emitter=None):
    target_str = target_date.strftime('%m/%d/%Y')

    # Bulk invoice — filter to target date
    bulk = pd.read_excel(bulk_path)
    bulk['Order #'] = pd.to_numeric(bulk['Order #'], errors='coerce')

    if 'ShipDate' in bulk.columns:
        bulk['ShipDate'] = bulk['ShipDate'].astype(str).str.strip()
        bulk_today = bulk[bulk['ShipDate'] == target_str]
        if bulk_today.empty:
            # Try M/D/YYYY without leading zeros (Windows strftime)
            try:
                alt = target_date.strftime('%-m/%-d/%Y')
            except ValueError:
                alt = target_str
            bulk_today = bulk[bulk['ShipDate'] == alt]
        bulk = bulk_today
        if emitter: emitter.log(f"Bulk invoice rows for {target_str}: {len(bulk)}")

    truck_map    = {}
    order_trucks = defaultdict(set)
    customer_map = {}
    dtype_map    = {}
    address_map  = {}
    bct_map      = {}   # billing customer type — multifamily detection
    piece_count  = defaultdict(int)

    for _, row in bulk.iterrows():
        if pd.isna(row['Order #']) or pd.isna(row.get('Truck')):
            continue
        oid       = int(row['Order #'])
        truck     = normalize_truck(str(row['Truck']).strip())
        model_key = str(row.get('Model Number') or '').strip().upper()
        order_trucks[oid].add(truck)
        if is_model(model_key):
            try:
                qty = int(row.get('Qty') or 1)
                qty = max(1, qty)  # ignore negatives
            except (ValueError, TypeError):
                qty = 1
            piece_count[oid] += qty
        if oid not in truck_map:
            truck_map[oid]       = truck
            customer_map[oid]    = str(row.get('Billing Customer') or '').strip()
            dtype_map[oid]       = str(row.get('Delivery/Pick-up Type') or '').strip()
            bct_map[oid]         = str(row.get('Billing Customer Type') or '').strip()
            addr_parts = [
                str(row.get('Shipping Address') or '').strip(),
                str(row.get('Shipping City')    or '').strip(),
                str(row.get('Shipping State')   or '').strip(),
                str(row.get('Shipping Zip')     or '').strip(),
            ]
            raw_addr = ', '.join(p for p in addr_parts if p)
            address_map[oid] = sanitize_address(raw_addr)

    # Orders detail — filtered to target date
    od = pd.read_csv(orders_path)
    od['Order #'] = pd.to_numeric(od['Order #'], errors='coerce')
    for col in ['SalePrice', 'Cost']:
        od[col] = od[col].astype(str).str.replace('[$,]', '', regex=True)
        od[col] = pd.to_numeric(od[col], errors='coerce').fillna(0)

    od_lines = defaultdict(list)
    for _, row in od.iterrows():
        if pd.isna(row['Order #']): continue
        oid = int(row['Order #'])
        if oid not in truck_map: continue
        od_lines[oid].append({
            'model':       str(row.get('Model Number') or '').strip(),
            'description': str(row.get('Description') or '').strip(),
            'qty':         int(row['Qty']) if pd.notna(row.get('Qty')) else 1,
            'sale_price':  float(row.get('SalePrice') or 0),
            'cost':        float(row.get('Cost') or 0),
            'est_date':    str(row.get('Est. Delivery') or '').strip(),
        })

    od_filtered = {}
    for oid, lines in od_lines.items():
        dated = [l for l in lines if l['est_date'] == target_str]
        od_filtered[oid] = dated if dated else lines

    # Detect storage release orders
    storage_orders = set()
    for oid, lines in od_filtered.items():
        for l in lines:
            if (l['model'].upper() == 'MEMO'
                    and 'DELIVER STORAGE ORDER' in l.get('description', '').upper()):
                storage_orders.add(oid)
                break
    if storage_orders:
        if emitter: emitter.log(f'Storage release orders: {sorted(storage_orders)}')

    # Detect swap orders — same model appears with both negative and positive qty
    swap_orders = set()
    for oid, lines in od_filtered.items():
        model_signs = defaultdict(set)
        for l in lines:
            m = l['model'].upper()
            if m in EXCLUDE_MODELS or not is_model(m): continue
            if l['qty'] < 0:   model_signs[m].add('neg')
            elif l['qty'] > 0: model_signs[m].add('pos')
        if any('neg' in s and 'pos' in s for s in model_signs.values()):
            swap_orders.add(oid)
    if swap_orders:
        if emitter: emitter.log(f'RMA/swap orders detected: {sorted(swap_orders)}')

    # Serial inventory
    ser = pd.read_csv(serial_path)
    ser['Order #'] = pd.to_numeric(ser['Order #'], errors='coerce')
    ser_lookup = defaultdict(lambda: defaultdict(list))
    # Full model→serial index for fallback cost lookup on previously invoiced units
    ser_by_model = defaultdict(list)
    for _, row in ser.iterrows():
        if pd.isna(row['Order #']): continue
        oid = int(row['Order #'])
        m = str(row.get('Model') or '').strip().upper()
        entry = {
            'cost':         float(row.get('Cost') or 0),
            'serial':       str(row.get('Serial') or ''),
            'inventory_id': str(row.get('Inventory Id') or ''),
        }
        ser_lookup[oid][m].append(entry)
        ser_by_model[m].append(entry)

    if emitter: emitter.log(f'Orders loaded: {len(truck_map)}')
    multi = {oid for oid, trucks in order_trucks.items() if len(trucks) > 1}
    if multi:
        if emitter: emitter.log(f'Multi-truck orders: {sorted(multi)}')

    return (truck_map, order_trucks, customer_map, dtype_map,
            address_map, od_filtered, ser_lookup, ser_by_model,
            dict(piece_count), bct_map, storage_orders, swap_orders)


# ── Charge calculation ─────────────────────────────────────────────

def calculate_charges(truck_map, order_trucks, customer_map, dtype_map,
                      address_map, od_filtered, piece_count, bct_map,
                      crate_status, stop_order, diesel_price,
                      storage_orders, swap_orders,
                      emitter=None, errors=None, prompt_answers=None):
    """
    Calculates delivery charges for all stops.
    emitter, errors, prompt_answers are required for app mode.
    """
    def log(msg, level='info'):
        if emitter: emitter.log(msg, level)
        else: print(msg)
    charges = {}

    multitruck_orders = {
        oid for oid, trucks in order_trucks.items()
        if len(trucks) > 1 and
        any(is_own_truck(t) for t in trucks)
    }

    # ── HUB trucks ────────────────────────────────────────────────
    hub_orders = [
        (oid, truck_map[oid]) for oid in truck_map
        if is_hub_truck(truck_map[oid])
        and categorize(truck_map[oid], dtype_map.get(oid,'')) == 'Delivery'
    ]

    if hub_orders:
        log(f'HUB stops — mileage lookup ({len(hub_orders)} stops)')
        wh_coords = geocode(WAREHOUSE); time.sleep(0.3)

        for oid, truck in hub_orders:
            name    = customer_map.get(oid, str(oid))
            address = address_map.get(oid, '')
            miles, mileage_c, fuel_c = hub_mileage_charges(address, oid, name, wh_coords, emitter=emitter)
            charges[oid] = {
                'fuel': fuel_c, 'mileage': mileage_c, 'floor': 0.0, 'threshold': 0.0,
                'x001_cost': 0.0, 'piece_cost': 0.0, 'labor': 0.0, 'insurance': 0.0,
                'crew': 0, 'total_extra': round(fuel_c + mileage_c, 2),
            }

        # HUB multifamily
        hub_mf = [oid for oid, _ in hub_orders if is_multifamily_customer(bct_map.get(oid, ''))]
        if hub_mf:
            log(f'HUB multifamily orders: {hub_mf}')
            for oid in hub_mf:
                name  = customer_map.get(oid, str(oid))
                lines = od_filtered.get(oid, [])
                has_products = any(
                    not str(l['model']).upper().startswith('X')
                    for l in lines
                    if str(l['model']).upper() not in EXCLUDE_MODELS
                )
                log(f'Order {oid} — {name} (multifamily)')
                units = prompt_int(
                    f'mf_{oid}_units',
                    {'order': oid, 'customer': name,
                     'question': 'How many units being serviced today?',
                     'sub': 'Multifamily threshold is calculated per unit'},
                    prompt_answers, emitter, allow_zero=False
                )
                threshold = round(MF_THRESHOLD * units, 2)
                floor_c   = 0.0
                if has_products:
                    has_upper = prompt_int(
                        f'mf_{oid}_upper_floors',
                        {'order': oid, 'customer': name,
                         'question': 'Units on floor 4 or above?',
                         'sub':      'Enter 0 for none, or the number of upper floors involved'},
                        prompt_answers, emitter, allow_zero=True
                    )
                    if has_upper and has_upper > 0:
                        for f_idx in range(has_upper):
                            floor_num  = 4 + f_idx
                            multiplier = floor_num - 3
                            pieces = prompt_int(
                                f'mf_{oid}_floor_{floor_num}',
                                {'order': oid, 'customer': name,
                                 'question': f'Floor {floor_num} — how many pieces?',
                                 'sub': f'${FLOOR_CHARGE:.2f} × {multiplier} multiplier per piece'},
                                prompt_answers, emitter, allow_zero=True
                            )
                            charge  = round(pieces * FLOOR_CHARGE * multiplier, 2)
                            floor_c += charge
                            if pieces > 0:
                                log(f'Floor {floor_num}: {pieces} × ${FLOOR_CHARGE} × {multiplier} = ${charge:.2f}')
                ex = charges.get(oid, {'fuel':0,'mileage':0,'floor':0,'threshold':0,
                                       'x001_cost':0,'piece_cost':0,'labor':0,'insurance':0,
                                       'crew':0,'total_extra':0})
                ex['floor']       = floor_c
                ex['threshold']   = threshold
                ex['total_extra'] = round(ex['fuel'] + ex['mileage'] + floor_c + threshold, 2)
                charges[oid] = ex

    # ── Storage release orders — prompt for piece count ───────────────
    storage_delivery = storage_orders & set(truck_map.keys())
    if storage_delivery:
        log(f'Storage release orders — prompting for piece count')
        for oid in sorted(storage_delivery):
            name      = customer_map.get(oid, str(oid))
            crate_key = crate_status.get(str(oid), 'OUT OF BOX')
            crate_lbl = 'IN-BOX (crated)' if crate_key == 'IN-BOX' else 'OUT OF BOX (uncrated)'
            rate      = CRATED_PER_PIECE if crate_key == 'IN-BOX' else UNCRATED_PER_PIECE
            log(f'Order {oid} — {name}: {crate_lbl} (logarithmic curve)')
            pieces = prompt_int(
                f'storage_{oid}_pieces',
                {'order': oid, 'customer': name,
                 'question': 'How many pieces releasing from storage?',
                 'sub': f'{crate_lbl} — cost calculated on logarithmic curve'},
                prompt_answers, emitter, allow_zero=False
            )
            piece_count[oid] = pieces
            log(f'Order {oid}: {pieces} pieces → ${piece_cost(pieces, crate_key):.2f}')

    # ── Own fleet single-truck stops ───────────────────────────────────
    own_by_truck = defaultdict(list)
    for oid in truck_map:
        truck = truck_map[oid]
        if (is_own_truck(truck)
                and categorize(truck, dtype_map.get(oid,'')) == 'Delivery'
                and oid not in multitruck_orders):
            own_by_truck[truck].append(oid)

    if own_by_truck:
        log(f'Own fleet routing — diesel ${diesel_price:.3f}/gal')

        for truck, order_ids in sorted(own_by_truck.items()):
            route_stops = stop_order.get(truck, [])
            route_order = {oid: idx for idx, (_, oid, _cust, _addr) in enumerate(route_stops)}
            sorted_ids  = sorted(order_ids, key=lambda oid: route_order.get(oid, 9999))
            log(f'Truck {truck} — {len(sorted_ids)} stops')

            crew = prompt_int(
                f'crew_{truck}',
                {'label': f'Truck {truck}',
                 'question': f'How many crew on Truck {truck} today?',
                 'sub': 'Number of people on the truck (affects piece cost curve)'},
                prompt_answers, emitter, allow_zero=False
            )
            log(f'Truck {truck}: {crew}-man crew')

            stops_for_route = []
            route_addr = {oid: addr for _, oid, _cust, addr in route_stops}
            for oid in sorted_ids:
                name = customer_map.get(oid, str(oid))
                addr = route_addr.get(oid) or address_map.get(oid, '')
                if addr:
                    stops_for_route.append((oid, name, addr))
                else:
                    log(f'Order {oid} ({name}): no address — $0 fuel', 'warn')

            fuel_per_stop = own_route_fuel(stops_for_route, diesel_price, emitter=emitter)

            for oid in sorted_ids:
                name        = customer_map.get(oid, str(oid))
                fuel_c      = fuel_per_stop.get(oid, 0.0)
                pieces      = piece_count.get(oid, 0)
                crate_key   = crate_status.get(str(oid), 'OUT OF BOX')
                rate        = CRATED_PER_PIECE if crate_key == 'IN-BOX' else UNCRATED_PER_PIECE
                crate_label = 'crated' if crate_key == 'IN-BOX' else 'uncrated'
                piece_c     = piece_cost(pieces, crate_key, crew)
                total_c     = round(X001_COST + piece_c + fuel_c, 2)
                log(f'Order {oid} ({name}): {pieces}pc {crate_label} {crew}-man ${piece_c:.2f} + fuel ${fuel_c:.2f} = ${total_c:.2f}')
                charges[oid] = {
                    'fuel': fuel_c, 'mileage': 0.0, 'floor': 0.0, 'threshold': 0.0,
                    'x001_cost': X001_COST, 'piece_cost': piece_c,
                    'labor': 0.0, 'insurance': 0.0, 'crew': crew,
                    'total_extra': total_c,
                }

    # ── Own fleet multi-truck stops (day-rate) ─────────────────────────
    if multitruck_orders:
        log(f'Multi-truck stops — day rate ({len(multitruck_orders)} orders)')
        for oid in sorted(multitruck_orders):
            trucks   = sorted(order_trucks[oid])
            name     = customer_map.get(oid, str(oid))
            n_trucks = len(trucks)
            log(f'Order {oid} — {name}: trucks {", ".join(trucks)}')

            total_crew = prompt_int(
                f'multitruck_{oid}_crew',
                {'order': oid, 'customer': name,
                 'question': 'Total crew across all trucks?',
                 'sub': f'Trucks: {", ".join(trucks)} — enter combined headcount'},
                prompt_answers, emitter, allow_zero=False
            )
            hours = prompt_int(
                f'multitruck_{oid}_hours',
                {'order': oid, 'customer': name,
                 'question': 'Hours on site?',
                 'sub': f'{total_crew} crew × ${LABOR_RATE:.0f}/hr × hours'},
                prompt_answers, emitter, allow_zero=False
            )

            labor_c     = round(total_crew * LABOR_RATE * hours, 2)
            insurance_c = round(INSURANCE_PER_TRUCK * n_trucks, 2)
            x001_total  = round(X001_COST * n_trucks, 2)

            total_fuel = 0.0
            for truck in trucks:
                truck_stops = stop_order.get(truck, [])
                route_addr  = {o: addr for _, o, _cust, addr in truck_stops}
                route_oids  = [o for _, o, _c, _a in truck_stops
                               if is_own_truck(truck_map.get(o, ''))
                               and categorize(truck_map.get(o,''), dtype_map.get(o,'')) == 'Delivery']
                stops_for_route = [
                    (o, customer_map.get(o, str(o)),
                     route_addr.get(o) or address_map.get(o,''))
                    for o in route_oids
                    if route_addr.get(o) or address_map.get(o,'')
                ]
                fuel_map    = own_route_fuel(stops_for_route, diesel_price, emitter=emitter)
                total_fuel += fuel_map.get(oid, 0.0)

            total_fuel = round(total_fuel, 2)
            total_c    = round(x001_total + insurance_c + labor_c + total_fuel, 2)
            log(f'Order {oid}: X001 ${x001_total:.2f} + ins ${insurance_c:.2f} + labor ${labor_c:.2f} + fuel ${total_fuel:.2f} = ${total_c:.2f}')

            charges[oid] = {
                'fuel': total_fuel, 'mileage': 0.0, 'floor': 0.0, 'threshold': 0.0,
                'x001_cost': x001_total, 'piece_cost': 0.0,
                'labor': labor_c, 'insurance': insurance_c,
                'crew': total_crew, 'total_extra': total_c,
            }

    return charges


# ── Build report rows ──────────────────────────────────────────────

def build_rows(truck_map, customer_map, dtype_map, od_filtered,
               ser_lookup, ser_by_model, charges, stop_order,
               storage_orders, swap_orders):
    fin_rows = []; svc_rows = []; prd_rows = []

    # Build global stop sort key: (truck_sort, stop_num_within_truck)
    stop_num_map = {}
    for truck, stops in stop_order.items():
        for stop_num, oid, _cust, _addr in stops:
            stop_num_map[oid] = (TRUCK_SORT.get(truck, 99), stop_num)

    for order_id in sorted(truck_map.keys()):
        truck    = truck_map[order_id]
        dtype    = dtype_map.get(order_id, '')
        customer = customer_map.get(order_id, str(order_id))
        bucket   = categorize(truck, dtype)
        if order_id in storage_orders and bucket == 'Delivery':
            bucket = 'Storage Release'
        if order_id in swap_orders and bucket == 'Delivery':
            bucket = 'RMA'
        lines    = od_filtered.get(order_id, [])
        ch       = charges.get(order_id, {})

        serial_used = defaultdict(int)
        svc_sale = svc_cost = prd_sale = prd_cost = 0.0

        # Detect swap: product model with both negative and positive lines
        swap_models = set()
        if order_id in swap_orders:
            model_signs = defaultdict(set)
            for line in lines:
                m = line['model'].upper()
                if m in EXCLUDE_MODELS or not is_model(m): continue
                if line['qty'] < 0:   model_signs[m].add('neg')
                elif line['qty'] > 0: model_signs[m].add('pos')
            swap_models = {m for m, s in model_signs.items()
                           if 'neg' in s and 'pos' in s}

        for line in lines:
            model  = str(line['model'] or '').strip()
            m_up   = model.upper()
            qty    = int(line['qty']) if line['qty'] else 1
            sale_p = float(line['sale_price'] or 0)
            cost_p = float(line['cost'] or 0)

            if m_up in EXCLUDE_MODELS: continue

            # ── Stair / oversized surcharges ──────────────────────
            if is_stair(model):
                if qty < 0: continue
                stair_sale = sale_p * qty
                stair_cost = round(stair_sale * STAIR_COST_RATIO, 2)
                profit = round(stair_sale - stair_cost, 2)
                margin = round((profit / stair_sale * 100) if stair_sale else 0.0, 1)
                svc_sale += stair_sale; svc_cost += stair_cost
                svc_rows.append({
                    'order_number': str(order_id), 'customer': customer,
                    'bucket': bucket, 'truck': truck, 'code': model,
                    'qty': qty, 'cost': stair_cost, 'sale': round(stair_sale, 2),
                    'profit_$': profit, 'margin_%': margin,
                })
                continue

            # ── X-code services ───────────────────────────────────
            if is_service(model):
                if qty < 0: continue
                sale   = sale_p
                cost   = XCODE_COST.get(m_up, cost_p)
                profit = round(sale - cost, 2)
                margin = round((profit / sale * 100) if sale else 0.0, 1)
                svc_sale += sale; svc_cost += cost
                svc_rows.append({
                    'order_number': str(order_id), 'customer': customer,
                    'bucket': bucket, 'truck': truck, 'code': model,
                    'qty': 1, 'cost': round(cost,2), 'sale': round(sale,2),
                    'profit_$': profit, 'margin_%': margin,
                })
                continue

            # ── Accessories (WATERLINE, DW KIT, etc.) ────────────
            if is_accessory(model):
                if qty < 0: continue
                acc_sale = sale_p * qty
                acc_cost = cost_p * qty
                profit = round(acc_sale - acc_cost, 2)
                margin = round((profit / acc_sale * 100) if acc_sale else 0.0, 1)
                prd_sale += acc_sale; prd_cost += acc_cost
                prd_rows.append({
                    'order_number': str(order_id), 'customer': customer,
                    'bucket': bucket, 'truck': truck, 'model': model,
                    'qty': qty, 'cost': round(acc_cost, 2), 'sale': round(acc_sale, 2),
                    'profit_$': profit, 'margin_%': margin, 'inventory_id': '',
                })
                continue

            # ── Swap: negative return line → RMA loss row ────────
            if m_up in swap_models and qty < 0:
                ser_list = ser_lookup[order_id].get(m_up, [])
                idx = serial_used[m_up]
                if idx < len(ser_list):
                    ret_cost     = float(ser_list[idx]['cost'] or 0)
                    ret_inv_id   = ser_list[idx]['inventory_id']
                    serial_used[m_up] += 1
                else:
                    history  = ser_by_model.get(m_up, [])
                    hist_idx = serial_used.get(m_up + '_HIST', 0)
                    if hist_idx < len(history):
                        ret_cost   = float(history[hist_idx]['cost'] or 0)
                        ret_inv_id = history[hist_idx]['inventory_id']
                        serial_used[m_up + '_HIST'] = hist_idx + 1
                    else:
                        ret_cost   = abs(cost_p)
                        ret_inv_id = ''
                # Revenue was already recognized on original delivery date.
                # Return row: sale=$0, cost=original cost paid → pure loss entry.
                # Vendor credit pending but not yet posted — show true cash-out position.
                ret_profit = round(0.0 - ret_cost, 2)
                prd_cost  += ret_cost
                prd_rows.append({
                    'order_number': str(order_id), 'customer': customer,
                    'bucket': bucket, 'truck': truck, 'model': f'RMA-RETURN:{model}',
                    'qty': abs(qty), 'cost': round(ret_cost, 2), 'sale': 0.0,
                    'profit_$': ret_profit, 'margin_%': 0.0,
                    'inventory_id': ret_inv_id,
                })
                continue

            # ── Standard product ──────────────────────────────────
            if qty < 0: continue

            ser_list = ser_lookup[order_id].get(m_up, [])
            idx = serial_used[m_up]
            if idx < len(ser_list):
                actual_cost  = float(ser_list[idx]['cost'] or 0)
                inventory_id = ser_list[idx]['inventory_id']
                has_serial   = True
                serial_used[m_up] += 1
            else:
                # Fallback: search full serial history by model for previously
                # invoiced units (scraper may have exported open items only)
                history  = ser_by_model.get(m_up, [])
                hist_idx = serial_used.get(m_up + '_HIST', 0)
                if hist_idx < len(history):
                    actual_cost  = float(history[hist_idx]['cost'] or 0)
                    inventory_id = history[hist_idx]['inventory_id']
                    serial_used[m_up + '_HIST'] = hist_idx + 1
                    has_serial   = True
                    if emitter: emitter.log(f'Order {order_id} {m_up}: cost from serial history (${actual_cost:.2f})')
                else:
                    actual_cost  = cost_p
                    inventory_id = ''
                    has_serial   = False

            effective_qty = 1 if has_serial else qty
            sale   = sale_p * effective_qty
            cost   = actual_cost * effective_qty
            profit = round(sale - cost, 2)
            margin = round((profit / sale * 100) if sale else 0.0, 1)
            prd_sale += sale; prd_cost += cost
            prd_rows.append({
                'order_number': str(order_id), 'customer': customer,
                'bucket': bucket, 'truck': truck, 'model': model,
                'qty': effective_qty, 'cost': round(cost, 2), 'sale': round(sale, 2),
                'profit_$': profit, 'margin_%': margin,
                'inventory_id': inventory_id,
            })

        def add_svc_cost(code, val):
            nonlocal svc_cost
            if val and val > 0:
                svc_cost += val
                svc_rows.append({
                    'order_number': str(order_id), 'customer': customer,
                    'bucket': bucket, 'truck': truck, 'code': code,
                    'qty': 1, 'cost': round(val,2), 'sale': 0.0,
                    'profit_$': round(-val,2), 'margin_%': 0.0,
                })

        add_svc_cost('X001-OVERHEAD', ch.get('x001_cost', 0))
        add_svc_cost('PIECES',        ch.get('piece_cost', 0))
        add_svc_cost('LABOR',         ch.get('labor', 0))
        add_svc_cost('INSURANCE',     ch.get('insurance', 0))
        add_svc_cost('FUEL',          ch.get('fuel', 0))
        add_svc_cost('MILEAGE',       ch.get('mileage', 0))
        add_svc_cost('FLOOR',         ch.get('floor', 0))
        add_svc_cost('THRESHOLD',     ch.get('threshold', 0))

        total_sale = round(prd_sale + svc_sale, 2)
        total_cost = round(prd_cost + svc_cost, 2)
        profit_t   = round(total_sale - total_cost, 2)
        margin_t   = round((profit_t / total_sale * 100) if total_sale else 0.0, 1)
        prd_profit = round(prd_sale - prd_cost, 2)
        prd_margin = round((prd_profit / prd_sale * 100) if prd_sale else 0.0, 1)
        svc_profit = round(svc_sale - svc_cost, 2)
        svc_margin = round((svc_profit / svc_sale * 100) if svc_sale else 0.0, 1)

        sort_key = stop_num_map.get(order_id, (TRUCK_SORT.get(truck, 99), 9999))

        fin_rows.append({
            'order_number': str(order_id), 'customer': customer,
            'city': '', 'bucket': bucket, 'truck': truck,
            'sort_key': sort_key,
            'total_sale': total_sale, 'total_cost': total_cost,
            'profit_$': profit_t, 'margin_%': margin_t,
            'svc_sale': round(svc_sale,2), 'svc_cost': round(svc_cost,2),
            'svc_profit_$': svc_profit, 'svc_margin_%': svc_margin,
            'prd_sale': round(prd_sale,2), 'prd_cost': round(prd_cost,2),
            'prd_profit_$': prd_profit, 'prd_margin_%': prd_margin,
            'balance_owed': 0.0,
            'crew': ch.get('crew', 0),
        })

    fin_rows.sort(key=lambda r: r['sort_key'])
    svc_rows.sort(key=lambda r: (
        stop_num_map.get(int(r['order_number']), (TRUCK_SORT.get(r['truck'],99),9999)),
    ))
    prd_rows.sort(key=lambda r: (
        stop_num_map.get(int(r['order_number']), (TRUCK_SORT.get(r['truck'],99),9999)),
    ))
    return fin_rows, svc_rows, prd_rows


# ── Write Excel ────────────────────────────────────────────────────

def write_excel(fin_rows, svc_rows, prd_rows, delivery_date):
    wb = Workbook(); wb.remove(wb.active)

    all_sale   = sum(r['total_sale'] for r in fin_rows)
    all_cost   = sum(r['total_cost'] for r in fin_rows)
    all_profit = round(all_sale - all_cost, 2)
    all_margin = round((all_profit / all_sale * 100) if all_sale else 0.0, 1)

    # Summary
    ws = wb.create_sheet("Summary")
    for col, w in zip("ABCDE", [32,16,16,14,12]):
        ws.column_dimensions[col].width = w
    r = 1
    c = ws.cell(row=r, column=1,
                value=f"Financial Summary — {delivery_date.strftime('%Y-%m-%d')} | All Trucks")
    c.font = bold(13); r += 2
    for ci, lbl in enumerate(["Bucket","Total Sale","Total Cost","Profit $","Margin %"],1):
        c = ws.cell(row=r, column=ci, value=lbl)
        c.font = bold(); c.fill = PatternFill("solid",fgColor=GRAY); c.border = BORDER
    r += 1
    for bucket in ["Delivery","Will Call","Storage Release","RMA"]:
        rows = [x for x in fin_rows if x['bucket']==bucket]
        if not rows: continue
        sale=sum(x['total_sale'] for x in rows)
        cost=sum(x['total_cost'] for x in rows)
        profit=round(sale-cost,2)
        margin=round((profit/sale*100) if sale else 0.0,1)
        for ci,(v,f) in enumerate(zip([bucket,sale,cost,profit,margin],
                                      ["text","dollar","dollar","dollar","pct"]),1):
            fmt_cell(ws.cell(row=r,column=ci,value=v),f)
        r+=1
    for ci,(v,f) in enumerate(zip(["TOTAL",all_sale,all_cost,all_profit,all_margin],
                                   ["text","dollar","dollar","dollar","pct"]),1):
        c=ws.cell(row=r,column=ci,value=v); fmt_cell(c,f); c.font=bold()

    # By Stop
    ws2 = wb.create_sheet("By Stop")
    hdrs = [
        ("Order #",12),("Customer",26),("City",14),("Bucket",12),("Truck",10),("Crew",6),
        ("Total Sale",13),("Total Cost",13),("Profit $",13),("Margin %",10),
        ("Svc Sale",13),("Svc Cost",13),("Svc Profit",13),("Svc Margin %",12),
        ("Prd Sale",13),("Prd Cost",13),("Prd Profit",13),("Prd Margin %",12),
        ("Balance Owed",13),
    ]
    write_headers(ws2, hdrs)
    FMTS = ["text","text","text","text","text","int",
            "dollar","dollar","dollar","pct",
            "dollar","dollar","dollar","pct",
            "dollar","dollar","dollar","pct","dollar"]
    BGS  = [None,None,None,None,None,None,
            BLUE,BLUE,BLUE,BLUE,
            ORANGE,ORANGE,ORANGE,ORANGE,
            GREEN,GREEN,GREEN,GREEN,None]
    for ri,x in enumerate(fin_rows,2):
        bf  = BUCKET_FILL.get(x['bucket'])
        vals= [x['order_number'],x['customer'],x['city'],x['bucket'],x['truck'],x['crew'],
               x['total_sale'],x['total_cost'],x['profit_$'],x['margin_%'],
               x['svc_sale'],x['svc_cost'],x['svc_profit_$'],x['svc_margin_%'],
               x['prd_sale'],x['prd_cost'],x['prd_profit_$'],x['prd_margin_%'],
               x['balance_owed']]
        for ci,(v,f,bg) in enumerate(zip(vals,FMTS,BGS),1):
            c=ws2.cell(row=ri,column=ci,value=v); fmt_cell(c,f,bg)
            if bf and not bg: c.fill=bf
    tr=len(fin_rows)+2
    ws2.cell(row=tr,column=1,value="TOTAL").font=bold()
    for ci,key,f in [(7,'total_sale','dollar'),(8,'total_cost','dollar'),(9,'profit_$','dollar'),
                      (11,'svc_sale','dollar'),(12,'svc_cost','dollar'),(13,'svc_profit_$','dollar'),
                      (15,'prd_sale','dollar'),(16,'prd_cost','dollar'),(17,'prd_profit_$','dollar')]:
        v=round(sum(x[key] for x in fin_rows),2)
        c=ws2.cell(row=tr,column=ci,value=v); fmt_cell(c,f); c.font=bold()
    for ci,nk,dk in [(10,'profit_$','total_sale'),(14,'svc_profit_$','svc_sale'),
                      (18,'prd_profit_$','prd_sale')]:
        n=sum(x[nk] for x in fin_rows); d=sum(x[dk] for x in fin_rows)
        v=round((n/d*100) if d else 0.0,1)
        c=ws2.cell(row=tr,column=ci,value=v); fmt_cell(c,'pct'); c.font=bold()

    # Service Detail
    ws3=wb.create_sheet("Service Detail")
    write_headers(ws3,[("Order #",12),("Customer",26),("Bucket",12),("Truck",10),
                        ("Code",16),("Qty",6),("Cost",13),("Sale",13),("Profit $",13),("Margin %",10)])
    for ri,x in enumerate(svc_rows,2):
        bf=BUCKET_FILL.get(x['bucket'])
        for ci,(v,f) in enumerate(zip(
            [x['order_number'],x['customer'],x['bucket'],x['truck'],
             x['code'],x['qty'],x['cost'],x['sale'],x['profit_$'],x['margin_%']],
            ["text","text","text","text","text","int","dollar","dollar","dollar","pct"]),1):
            c=ws3.cell(row=ri,column=ci,value=v); fmt_cell(c,f)
            if bf: c.fill=bf
    # Totals row
    tr3 = len(svc_rows) + 2
    ws3.cell(row=tr3, column=1, value="TOTAL").font = bold()
    t_qty  = sum(x['qty']       for x in svc_rows)
    t_cost = round(sum(x['cost']  for x in svc_rows), 2)
    t_sale = round(sum(x['sale']  for x in svc_rows), 2)
    t_prof = round(t_sale - t_cost, 2)
    t_marg = round((t_prof / t_sale * 100) if t_sale else 0.0, 1)
    for ci,(v,f) in enumerate(zip([t_qty, t_cost, t_sale, t_prof, t_marg],
                                   ["int","dollar","dollar","dollar","pct"]), 6):
        c = ws3.cell(row=tr3, column=ci, value=v); fmt_cell(c, f); c.font = bold()

    # Product Detail
    ws4=wb.create_sheet("Product Detail")
    write_headers(ws4,[("Order #",12),("Customer",26),("Bucket",12),("Truck",10),
                        ("Model",18),("Qty",6),("Cost",13),("Sale",13),
                        ("Profit $",13),("Margin %",10),("Inventory ID",16)])
    for ri,x in enumerate(prd_rows,2):
        bf=BUCKET_FILL.get(x['bucket'])
        for ci,(v,f) in enumerate(zip(
            [x['order_number'],x['customer'],x['bucket'],x['truck'],
             x['model'],x['qty'],x['cost'],x['sale'],x['profit_$'],x['margin_%'],
             x['inventory_id']],
            ["text","text","text","text","text","int","dollar","dollar","dollar","pct","text"]),1):
            c=ws4.cell(row=ri,column=ci,value=v); fmt_cell(c,f)
            if bf: c.fill=bf
    # Totals row
    tr4 = len(prd_rows) + 2
    ws4.cell(row=tr4, column=1, value="TOTAL").font = bold()
    p_qty  = sum(x['qty']      for x in prd_rows)
    p_cost = round(sum(x['cost'] for x in prd_rows), 2)
    p_sale = round(sum(x['sale'] for x in prd_rows), 2)
    p_prof = round(p_sale - p_cost, 2)
    p_marg = round((p_prof / p_sale * 100) if p_sale else 0.0, 1)
    for ci,(v,f) in enumerate(zip([p_qty, p_cost, p_sale, p_prof, p_marg],
                                   ["int","dollar","dollar","dollar","pct"]), 6):
        c = ws4.cell(row=tr4, column=ci, value=v); fmt_cell(c, f); c.font = bold()

    month_name = delivery_date.strftime('%B').upper()
    year_str   = delivery_date.strftime('%Y')
    out_dir    = os.path.join(EXPORT_BASE, year_str, month_name)
    os.makedirs(out_dir, exist_ok=True)
    filepath   = os.path.join(out_dir, f"DeliveryReport_{delivery_date.strftime('%m%d%Y')}.xlsx")
    wb.save(filepath)
    return filepath


def _emit_summary(fin_rows, emitter):
    """Sends the summary table to the live log feed."""
    all_sale   = sum(r['total_sale'] for r in fin_rows)
    all_cost   = sum(r['total_cost'] for r in fin_rows)
    all_profit = round(all_sale - all_cost, 2)
    all_margin = round((all_profit / all_sale * 100) if all_sale else 0.0, 1)
    emitter.log('─' * 60, 'header')
    for r in fin_rows:
        emitter.log(
            f"{r['order_number']:<8} {str(r['customer'])[:22]:<23} "
            f"{r['bucket']:<14} {str(r['truck']):<8} "
            f"${r['total_sale']:>9,.0f}  {r['margin_%']:>5.1f}%"
        )
    emitter.log('─' * 60, 'header')
    emitter.log(
        f"{'TOTAL':<8} {len(fin_rows):>3} stops{'':<17} "
        f"${all_sale:>9,.0f}  {all_margin:>5.1f}%",
        'good'
    )


def _build_map_data(fin_rows, stop_order, address_map, customer_map, charges, emitter=None):
    """
    Builds the route data structure for the Google Maps panel.
    Geocodes stop addresses server-side so the map renders immediately.
    Returns a dict ready for emitter.map_data().
    """
    # Geocode cache — avoid duplicate API calls for same address
    _geo_cache = {}

    def geo(addr):
        if not addr: return None, None
        if addr in _geo_cache: return _geo_cache[addr]
        coords = geocode(addr, emitter)
        time.sleep(0.15)   # gentle rate limit
        if coords:
            # coords is [lng, lat] — swap to lat, lng for Maps JS
            result = (coords[1], coords[0])
        else:
            result = (None, None)
        _geo_cache[addr] = result
        return result

    trucks = {}
    for r in fin_rows:
        t = r['truck']
        if t not in trucks:
            trucks[t] = {'id': t, 'stops': [], 'sale': 0, 'cost': 0}
        trucks[t]['sale'] += r['total_sale']
        trucks[t]['cost'] += r['total_cost']

    order_fin = {r['order_number']: r for r in fin_rows}

    if emitter: emitter.log('Geocoding stop addresses for map…')

    # Build stops from route sheet (has correct stop order)
    for truck, stops in stop_order.items():
        if truck not in trucks: continue
        for stop_num, oid, customer, addr in stops:
            r   = order_fin.get(str(oid), {})
            lat, lng = geo(addr or address_map.get(oid, ''))
            trucks[truck]['stops'].append({
                'stop_num': stop_num,
                'order':    oid,
                'customer': customer or r.get('customer', str(oid)),
                'address':  addr or address_map.get(oid, ''),
                'sale':     r.get('total_sale', 0),
                'cost':     r.get('total_cost', 0),
                'margin':   r.get('margin_%', 0),
                'lat':      lat,
                'lng':      lng,
            })

    # For trucks not in route sheet, build from fin_rows
    for truck_id, data in trucks.items():
        if not data['stops']:
            truck_rows = [r for r in fin_rows if r['truck'] == truck_id]
            for i, r in enumerate(truck_rows, 1):
                oid  = int(r['order_number'])
                addr = address_map.get(oid, '')
                lat, lng = geo(addr)
                data['stops'].append({
                    'stop_num': i,
                    'order':    oid,
                    'customer': r['customer'],
                    'address':  addr,
                    'sale':     r['total_sale'],
                    'cost':     r['total_cost'],
                    'margin':   r['margin_%'],
                    'lat':      lat,
                    'lng':      lng,
                })

    # Add margin to truck summaries
    for t, d in trucks.items():
        d['margin'] = round(((d['sale'] - d['cost']) / d['sale'] * 100)
                            if d['sale'] else 0, 1)

    geocoded = sum(1 for d in trucks.values()
                   for s in d['stops'] if s['lat'] is not None)
    total_stops = sum(len(d['stops']) for d in trucks.values())
    if emitter: emitter.log(f'Map geocoded: {geocoded}/{total_stops} stops', 'good')

    return {
        'trucks':         list(trucks.values()),
        'google_api_key': os.getenv("GOOGLE_API_KEY", GOOGLE_API_KEY),
    }


# ── Main (app version) ─────────────────────────────────────────────

def run(bulk_path, serial_path, orders_path,
        route_sheet_path, delivery_date,
        emitter, errors, prompt_answers=None,
        diesel_price=3.80):
    """
    App entry point — called by pipeline.py.

    Args:
        bulk_path, serial_path, orders_path, route_sheet_path: file paths
        delivery_date:  datetime
        emitter:        ProgressEmitter
        errors:         ErrorRegistry
        prompt_answers: shared dict for UI prompt answers

    Returns:
        (filepath: str, fin_rows: list)
    """
    prompt_answers = prompt_answers if prompt_answers is not None else {}

    emitter.running('charges', 'Loading export files')

    # Re-read env vars (may have been updated after import)
    global ORS_API_KEY, GOOGLE_API_KEY, MONDAY_TOKEN
    ORS_API_KEY    = os.getenv("ORS_API_KEY", "")
    GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")
    MONDAY_TOKEN   = os.getenv("MONDAY_API_TOKEN", "")

    (truck_map, order_trucks, customer_map, dtype_map,
     address_map, od_filtered, ser_lookup, ser_by_model,
     piece_count, bct_map, storage_orders, swap_orders) = load_files(
        bulk_path, serial_path, orders_path, delivery_date, emitter=emitter
    )

    if not truck_map:
        errors.fatal('build', 'NO_ORDERS',
                     f'No delivery orders found for {delivery_date.strftime("%m/%d/%Y")}. '
                     'Check the bulk invoice date filter.')

    # Route sheet
    stop_order = {}
    if route_sheet_path and os.path.exists(route_sheet_path):
        emitter.log('Parsing route sheet…')
        stop_order = parse_route_sheet(route_sheet_path, emitter=emitter)
    else:
        errors.warn('route', 'NO_ROUTE_SHEET',
                    'No route sheet — stops sorted by truck then order number')

    # Monday crate status
    emitter.running('monday', 'Fetching crate status from Monday')
    own_order_ids = [
        oid for oid in truck_map
        if is_own_truck(truck_map[oid])
        and categorize(truck_map[oid], dtype_map.get(oid, '')) == 'Delivery'
    ]
    crate_status = get_crate_status(own_order_ids, emitter=emitter)
    emitter.done('monday', f'Crate status: {len(crate_status)} orders')

    emitter.log(f'Diesel price: ${diesel_price:.3f}/gal')

    # Charges
    emitter.running('charges', 'Calculating delivery costs')
    charges = calculate_charges(
        truck_map, order_trucks, customer_map, dtype_map,
        address_map, od_filtered, piece_count, bct_map,
        crate_status, stop_order, diesel_price,
        storage_orders, swap_orders,
        emitter=emitter, errors=errors, prompt_answers=prompt_answers
    )
    emitter.done('charges', 'Delivery costs calculated')

    # Build rows
    emitter.running('build', 'Building report rows')
    fin_rows, svc_rows, prd_rows = build_rows(
        truck_map, customer_map, dtype_map, od_filtered,
        ser_lookup, ser_by_model, charges, stop_order,
        storage_orders, swap_orders
    )
    emitter.done('build', f'Built {len(fin_rows)} stops')

    _emit_summary(fin_rows, emitter)

    # Write Excel
    emitter.running('excel', 'Writing Excel report')
    filepath = write_excel(fin_rows, svc_rows, prd_rows, delivery_date)
    emitter.done('excel', f'Saved: {os.path.basename(filepath)}')

    # Emit map data
    map_data = _build_map_data(fin_rows, stop_order, address_map, customer_map, charges, emitter=emitter)
    emitter.log(f'Map data: {len(map_data.get("trucks",[]))} trucks, key={bool(map_data.get("google_api_key"))}')
    emitter.map_data(map_data)

    # Serialize fin_rows for UI (only JSON-safe types)
    serializable = [
        {k: v for k, v in r.items() if k != 'sort_key'}
        for r in fin_rows
    ]

    return filepath, serializable
